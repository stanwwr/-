import numpy as np
import scipy.sparse as sp
from decimal import Decimal
import openpyxl as xl
def construct_matrix(filename, m1, m2):
    if filename=="DV(exp).txt":
        matrix = np.zeros((m1, m2), dtype = np.float)
    else:
        matrix = np.zeros((m1, m2), dtype = np.float)
    #adj_matrix=np.zeros((m1, m2), dtype = np.int)
    #print(matrix)
    with open(filename, "r") as f:
        for line in f:
            entry = [i for i in line.split()]
            entry[0]=int(entry[0])
            entry[1]=int(entry[1])
            if filename=="DV(exp).txt":
                entry[2]=float(entry[2])
            else:
                entry[2]=int(entry[2])
            #print(entry[0])
            #print(entry[1])
            #print(entry[2])
            matrix[entry[0]-1][entry[1]-1] = entry[2]
            #adj_matrix[entry[0]-1][entry[1]-1]=1
    #out = sp.csc_matrix(matrix)
    #out = sp.csc_matrix(matrix)
    return matrix
def Get_expscore(xlsx):
    wb = xl.load_workbook(xlsx)
    # 仅支持单sheet
    sheet = wb.active
    i = 2
    s=0
    vector=[]
    while(sheet["C%d" % i].value != None):
        t=float(sheet["C%d" % i].value)
        vector.append(t)
        i += 1
        s+=t
    return vector
def Get_impscore(xlsx):
    wb = xl.load_workbook(xlsx)
    # 仅支持单sheet
    sheet = wb.active
    i = 2
    s=0
    vector=[]
    while(sheet["E%d" % i].value != None):
        t=float(sheet["E%d" % i].value)
        vector.append(t)
        i += 1
        s+=t
    return vector
def Get_basescore(xlsx):
    wb = xl.load_workbook(xlsx)
    # 仅支持单sheet
    sheet = wb.active
    i = 2
    s=0
    vector=[]
    while(sheet["C%d" % i].value != None):
        t=float(sheet["C%d" % i].value)
        vector.append(t)
        i += 1
        s+=t
    return vector
def Nor_vector(vec1):
    s=float(0)
    for i in range(len(vec1)):
        s+=vec1[i]
    for i in range(len(vec1)):
        vec1[i]=vec1[i]/s
    return vec1
def Process_to_output(vec):
    lis=[]
    for i in range(len(vec)):
        lis.append((vec[i],i+1))
    #lis.sort(reverse=True)
    return lis
    
def Nor_matrix_col(matrix):
    row=matrix.shape[0]
    col=matrix.shape[1]
    #print(row,col)
    vect=[0 for i in range(col)]
    for i in range(col):
        for j in range(row):
            vect[i]+=matrix[j][i]
        #print(vect[i])
    for i in range(col):
        for j in range(row):
            #print(matrix[j][i],vect[i])
            if vect[i]==0:
                matrix[j][i]=0
            else:
                matrix[j][i]=matrix[j][i]/vect[i]
            #print(matrix[j][i])
    return matrix

def Get_order(p_imp, cvss_imp, risk_score, cvss_score, filename):
    a=[0 for i in range(len(p_imp))]
    b=[0 for i in range(len(cvss_imp))]
    c=[0 for i in range(len(risk_score))]
    d=[0 for i in range(len(cvss_score))]
    for i in range(len(p_imp)):
        a[p_imp[i][1]-1]=i
    for i in range(len(cvss_imp)):
        b[cvss_imp[i][1]-1]=i
    for i in range(len(cvss_score)):
        c[risk_score[i][1]-1]=i
    for i in range(len(cvss_score)):
        d[cvss_score[i][1]-1]=i

    f=open(filename,'w+')
    for i in range(len(a)):
        f.write(str(a[i]+1)+' ')
    f.write('\n')
    for i in range(len(b)):
        f.write(str(b[i]+1)+' ')
    f.write('\n')
    for i in range(len(c)):
        f.write(str(c[i]+1)+' ')    
    f.write('\n')
    for i in range(len(d)):
        f.write(str(d[i]+1)+' ')
    f.write('\n')
def PathSim(x, m, k, num):
    """ # this is significantly slower
    candidates = set()
    for i in matrix.getrow(x-1).indices:
        for j in matrix.getrow(i-1).indices:
            candidates.add(j)
    """     
    out = []
    
    for i in range(num):
        value = 2.0 * m[x-1,i] / (m[x-1,x-1] + m[i,i])
        out.append((value, i+1))
        
    #out.sort()
    #out.reverse()
    return out[:k]
def pagerank(mat,v,v0,a):
    row=mat.shape[0]
    #col1=trans1.shape[1]
    DD=Nor_matrix_col(mat)
    print(DD)
    v1=v[:]
    #distance=v[:]
    #print(distance)
    for k in range(100):
        #print(distance)
        #print(v)
        #flag=0
        for i in range(row):
            temp0=0
            for j in range(row):
                temp0+=DD[i][j]*v[j]
            v1[i]=a*temp0+(1-a)*v0[i]
            #v1[i]=temp0
        
        v1=Nor_vector(v1)
        v=v1[:]
        #print("try1111111",r_V_exp)
        final=Process_to_output(v)
        
       
        
        final.sort()
        final.reverse()
        print(k,":::::")
        print("////////")
        print(final)
        print("////////////////////////")
def main():
    
    W_DV=np.zeros((8,13), dtype=np.float)
    W_DD=np.zeros((7,7), dtype=np.float)
    W_VA=np.zeros((13,6), dtype=np.float)


    W_DV[2][0]=6.4
    W_DV[2][1]=6.4
    W_DV[2][2]=6.9

    W_DV[3][3]=10.0
    W_DV[1][4]=6.9
    W_DV[1][5]=10.0
    W_DV[1][6]=6.4
    W_DV[4][7]=10.0
    W_DV[5][8]=10.0
    W_DV[5][9]=2.9
    W_DV[5][10]=6.4
    W_DV[6][11]=2.9
    W_DV[0][12]=10

    W_DV[7][0]=6.4
    W_DV[7][1]=6.4
    W_DV[7][2]=6.9
    W_DV[7][3]=10.0
    W_DV[7][4]=6.9
    W_DV[7][5]=10.0
    W_DV[7][6]=6.4
    W_DV[7][7]=10.0
    W_DV[7][8]=10.0
    W_DV[7][9]=2.9
    W_DV[7][10]=6.4
    W_DV[7][11]=2.9
    W_DV[7][12]=10
    # W_DD init
    W_DD[0][4]=1
    W_DD[1][2]=1
    W_DD[1][3]=1
    W_DD[2][1]=2
    W_DD[2][5]=2
    W_DD[2][4]=1
    W_DD[2][0]=1
    W_DD[2][6]=2
    W_DD[3][1]=2
    W_DD[3][5]=2
    W_DD[3][4]=1
    W_DD[3][0]=1
    W_DD[3][6]=2

    W_VA[0][0]=10
    W_VA[0][1]=10
    W_VA[0][2]=10
    W_VA[1][1]=10
    W_VA[1][2]=10
    W_VA[1][3]=10
    W_VA[2][1]=3.9
    W_VA[3][1]=8.6
    W_VA[3][2]=8.6
    W_VA[3][4]=8.6
    W_VA[4][1]=10.0
    W_VA[5][4]=10.0
    W_VA[5][1]=10.0
    W_VA[5][2]=10.0
    W_VA[6][5]=10.0
    W_VA[6][1]=10.0
    W_VA[6][2]=10.0
    W_VA[7][1]=8.0
    W_VA[7][2]=8.0
    W_VA[7][3]=8.0
    W_VA[8][1]=10
    W_VA[8][2]=10
    W_VA[8][4]=10
    W_VA[9][1]=10
    W_VA[10][1]=10
    W_VA[10][2]=10
    W_VA[10][3]=10
    W_VA[11][3]=10
    W_VA[12][3]=4.9
    W_VA[12][2]=4.9
    W_VA[12][1]=4.9

    
    W_AV=W_VA.T[:]
    W_VD=W_DV.T[:]
    W_DD=W_DD.T[:]
    W_AV = sp.csc_matrix(W_AV)
    W_VA = sp.csc_matrix(W_VA)
    W_DV = sp.csc_matrix(W_DV)
    W_VD = sp.csc_matrix(W_VD)
    final_result=[]
    DVD =(W_DV*W_VD).todense()
    DVAVD = (W_DV*W_VA*W_AV*W_VD).todense()
    result_DVD=PathSim(8,DVD,8,8)
    result_DVAVD=PathSim(8,DVAVD,8,8)
    initial_vec=[]
    for i in range(len(result_DVAVD)):
        final_result.append((result_DVAVD[i][0]+result_DVD[i][0],i+1))
    for i in range(len(final_result)-1):
        initial_vec.append(final_result[i][0])
    final_result.sort()
    final_result.reverse()
    initial_vec=Nor_vector(initial_vec)
    
    print('asdfasdfasdf')
    print(final_result)
    print('asdfasdfasdf')
    print(initial_vec)
    pagerank(W_DD,initial_vec,initial_vec,0.5)
    #pagerank(W_DV,W_DD,init_D,init_V,0.5)
    #r_V=Get_impscore("vulnerability-infor.xlsx")
    #r_V=[6.4,6.4,6.9,10.0,6.9,10.0,6.4,10.0,10.0,2.9,6.4,2.9,10.0]
    #r_V=[float(1/37) for i in range(37)]
    #r_V=Nor_vector(r_V)
    #print(r_V)
    #W_DV=Nor_matrix_col(W_DV)
    #W_VD=Nor_matrix_col(W_VD)
    #W_D=Nor_matrix_col(W_D)
    #print(W_D)
    #print("fuuuuuuuck",r_V_exp)
    #print(r_V,len(r_V),s_r_V)
    #f_conver_data=open('convergence_data.txt','w+')
 
        #print(r_V_exp)
    #f_conver_data.close()
    #r_V_exp=Get_expscore("VA.xlsx")
    #r_V_exp=Nor_vector(r_V_exp)
    #r_V_exp_final=Process_to_output(r_V_exp)
    #r_V_base=Process_to_output(Nor_vector(Get_basescore("vulnerability-infor.xlsx")))
    #print(r_V_exp,"///")
    '''score_final=[]
    for i in range(37):
        score_final.append((r_V_final[i][0]+r_V_exp_final[i][0],i+1))
    score_final.sort()
    score_final.reverse()
    r_V_final.sort()
    r_V_final.reverse()
    r_V_exp_final.sort()
    r_V_exp_final.reverse()
    r_V_base.sort()
    r_V_base.reverse()
    V_impscore=Process_to_output(Nor_vector(Get_impscore("vulnerability-infor.xlsx")))
    V_impscore.sort()
    V_impscore.reverse()
    print(r_V_exp)
    print(score_final)
    Get_order(r_V_final,V_impscore,score_final,r_V_base,'ord_figure.txt')

    f_table1_tex=open('table1_tex.txt','w+')
    for i in range(37):
        f_table1_tex.write(str(i+1)+' '+'&'+' '+str(round(r_V_final[i][0],6))+' '+'&'+' '+str(r_V_final[i][1])+' '+'&'+' '+str(round(V_impscore[i][0],6))+' '+'&'+' '+str(V_impscore[i][1])+' '+'&'+' '+str(round(score_final[i][0],6))+' '+'&'+' '+str(score_final[i][1])+' '+'&'+' '+str(round(r_V_base[i][0],6))+' '+'&'+' '+str(r_V_base[i][1])+' '+'\\'+'\\'+' '+'\hline'+'\n')
    f_table1_tex.close()'''


        

if __name__ =='__main__':
    main()  
